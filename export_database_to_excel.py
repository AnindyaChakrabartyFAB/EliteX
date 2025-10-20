#!/usr/bin/env python3
"""
Elite Database Export Script
Exports all core and app schema tables to well-structured Excel files
"""

import os
import sys
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
import logging
from datetime import datetime
from pathlib import Path
import numpy as np

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('database_export.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class EliteDatabaseExporter:
    """Export all Elite database tables to Excel files"""
    
    def __init__(self):
        """Initialize database connection"""
        load_dotenv()
        self.db_config = {
            "host": os.getenv("DB_HOST", "localhost"),
            "port": os.getenv("DB_PORT", "5432"),
            "database": os.getenv("DB_NAME", "elite_db"),
            "user": os.getenv("DB_USER", "postgres"),
            "password": os.getenv("DB_PASSWORD", "")
        }
        self.connection = None
        self.data_folder = Path("Data")
        self._connect()
    
    def _connect(self):
        """Establish database connection"""
        try:
            self.connection = psycopg2.connect(**self.db_config)
            logger.info("✅ Connected to Elite PostgreSQL database")
        except Exception as e:
            logger.error(f"❌ Database connection failed: {e}")
            raise
    
    def get_all_tables(self):
        """Get all tables from core and app schemas"""
        query = """
        SELECT 
            schemaname,
            tablename,
            tableowner
        FROM pg_tables 
        WHERE schemaname IN ('core', 'app')
        ORDER BY schemaname, tablename
        """
        
        try:
            with self.connection.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute(query)
                return cursor.fetchall()
        except Exception as e:
            logger.error(f"❌ Failed to get table list: {e}")
            return []
    
    def get_table_info(self, schema_name, table_name):
        """Get detailed table information including columns"""
        query = """
        SELECT 
            column_name,
            data_type,
            is_nullable,
            column_default,
            character_maximum_length,
            numeric_precision,
            numeric_scale
        FROM information_schema.columns 
        WHERE table_schema = %s AND table_name = %s
        ORDER BY ordinal_position
        """
        
        try:
            with self.connection.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute(query, (schema_name, table_name))
                return cursor.fetchall()
        except Exception as e:
            logger.error(f"❌ Failed to get table info for {schema_name}.{table_name}: {e}")
            return []
    
    def get_table_data(self, schema_name, table_name, limit=10000):
        """Get data from table with optional limit"""
        query = f"""
        SELECT * FROM {schema_name}.{table_name}
        LIMIT %s
        """
        
        try:
            with self.connection.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute(query, (limit,))
                data = cursor.fetchall()
                
                # Convert timezone-aware datetimes to timezone-naive for Excel compatibility
                for row in data:
                    for key, value in row.items():
                        if hasattr(value, 'tzinfo') and value.tzinfo is not None:
                            # Convert timezone-aware datetime to naive datetime
                            row[key] = value.replace(tzinfo=None)
                
                return data
        except Exception as e:
            logger.error(f"❌ Failed to get data from {schema_name}.{table_name}: {e}")
            return []
    
    def get_table_row_count(self, schema_name, table_name):
        """Get total row count for table"""
        query = f"""
        SELECT COUNT(*) as row_count FROM {schema_name}.{table_name}
        """
        
        try:
            with self.connection.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute(query)
                result = cursor.fetchone()
                return result['row_count'] if result else 0
        except Exception as e:
            logger.error(f"❌ Failed to get row count for {schema_name}.{table_name}: {e}")
            return 0
    
    def create_excel_file(self, schema_name, table_name, table_info, table_data, row_count):
        """Create well-structured Excel file for a table"""
        
        # Create filename
        filename = f"{schema_name}_{table_name}.xlsx"
        filepath = self.data_folder / filename
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # Sheet 1: Table Information
                info_data = []
                for col in table_info:
                    info_data.append({
                        'Column Name': col['column_name'],
                        'Data Type': col['data_type'],
                        'Nullable': col['is_nullable'],
                        'Default Value': col['column_default'] or '',
                        'Max Length': col['character_maximum_length'] or '',
                        'Numeric Precision': col['numeric_precision'] or '',
                        'Numeric Scale': col['numeric_scale'] or ''
                    })
                
                info_df = pd.DataFrame(info_data)
                info_df.to_excel(writer, sheet_name='Table_Info', index=False)
                
                # Sheet 2: Sample Data
                if table_data:
                    data_df = pd.DataFrame(table_data)
                    
                    # Convert timezone-aware datetime columns to timezone-naive
                    for col in data_df.columns:
                        if data_df[col].dtype == 'object':
                            # Check if column contains datetime objects
                            sample_val = data_df[col].dropna().iloc[0] if not data_df[col].dropna().empty else None
                            if sample_val is not None and hasattr(sample_val, 'tzinfo') and sample_val.tzinfo is not None:
                                # Convert timezone-aware to timezone-naive
                                data_df[col] = data_df[col].apply(
                                    lambda x: x.replace(tzinfo=None) if pd.notna(x) and hasattr(x, 'tzinfo') and x.tzinfo is not None else x
                                )
                    
                    data_df.to_excel(writer, sheet_name='Sample_Data', index=False)
                else:
                    # Create empty sheet with column headers
                    empty_df = pd.DataFrame(columns=[col['column_name'] for col in table_info])
                    empty_df.to_excel(writer, sheet_name='Sample_Data', index=False)
                
                # Sheet 3: Summary
                summary_data = {
                    'Property': [
                        'Schema Name',
                        'Table Name',
                        'Total Rows',
                        'Columns Count',
                        'Export Date',
                        'Sample Size',
                        'File Created'
                    ],
                    'Value': [
                        schema_name,
                        table_name,
                        row_count,
                        len(table_info),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        min(10000, row_count) if row_count > 0 else 0,
                        filename
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            logger.info(f"✅ Created Excel file: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to create Excel file {filename}: {e}")
            return False
    
    def format_excel_file(self, filepath):
        """Apply formatting to Excel file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = load_workbook(filepath)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Apply borders to all cells with data
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell.border = border
            
            wb.save(filepath)
            logger.info(f"✅ Applied formatting to {filepath.name}")
            
        except Exception as e:
            logger.error(f"❌ Failed to format Excel file {filepath}: {e}")
    
    def export_all_tables(self):
        """Export all tables to Excel files"""
        logger.info("🚀 Starting Elite Database Export")
        logger.info("=" * 60)
        
        # Ensure data folder exists
        self.data_folder.mkdir(exist_ok=True)
        
        # Get all tables
        tables = self.get_all_tables()
        if not tables:
            logger.error("❌ No tables found in core and app schemas")
            return
        
        logger.info(f"📊 Found {len(tables)} tables to export")
        
        # Export each table
        success_count = 0
        failed_count = 0
        
        for table in tables:
            schema_name = table['schemaname']
            table_name = table['tablename']
            
            logger.info(f"📋 Processing {schema_name}.{table_name}...")
            
            try:
                # Get table information
                table_info = self.get_table_info(schema_name, table_name)
                if not table_info:
                    logger.warning(f"⚠️  No column information found for {schema_name}.{table_name}")
                    continue
                
                # Get row count
                row_count = self.get_table_row_count(schema_name, table_name)
                logger.info(f"   📈 Row count: {row_count:,}")
                
                # Get sample data (limit to 10,000 rows for performance)
                table_data = self.get_table_data(schema_name, table_name, limit=10000)
                logger.info(f"   📊 Sample data rows: {len(table_data)}")
                
                # Create Excel file
                if self.create_excel_file(schema_name, table_name, table_info, table_data, row_count):
                    # Apply formatting
                    filename = f"{schema_name}_{table_name}.xlsx"
                    filepath = self.data_folder / filename
                    self.format_excel_file(filepath)
                    success_count += 1
                else:
                    failed_count += 1
                
            except Exception as e:
                logger.error(f"❌ Failed to export {schema_name}.{table_name}: {e}")
                failed_count += 1
        
        # Create summary report
        self.create_summary_report(tables, success_count, failed_count)
        
        logger.info("=" * 60)
        logger.info(f"✅ Export completed!")
        logger.info(f"   📊 Total tables: {len(tables)}")
        logger.info(f"   ✅ Successfully exported: {success_count}")
        logger.info(f"   ❌ Failed exports: {failed_count}")
        logger.info(f"   📁 Files saved to: {self.data_folder.absolute()}")
    
    def create_summary_report(self, tables, success_count, failed_count):
        """Create a summary report of all exported tables"""
        try:
            summary_data = []
            
            for table in tables:
                schema_name = table['schemaname']
                table_name = table['tablename']
                
                # Get row count
                row_count = self.get_table_row_count(schema_name, table_name)
                
                # Check if file exists
                filename = f"{schema_name}_{table_name}.xlsx"
                filepath = self.data_folder / filename
                file_exists = filepath.exists()
                
                summary_data.append({
                    'Schema': schema_name,
                    'Table Name': table_name,
                    'Row Count': row_count,
                    'Excel File': filename,
                    'File Created': 'Yes' if file_exists else 'No',
                    'Export Status': 'Success' if file_exists else 'Failed'
                })
            
            # Create summary DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # Save summary report
            summary_file = self.data_folder / "00_DATABASE_EXPORT_SUMMARY.xlsx"
            
            with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='Export_Summary', index=False)
                
                # Add statistics sheet
                stats_data = {
                    'Metric': [
                        'Total Tables',
                        'Successfully Exported',
                        'Failed Exports',
                        'Total Rows Across All Tables',
                        'Export Date',
                        'Export Duration'
                    ],
                    'Value': [
                        len(tables),
                        success_count,
                        failed_count,
                        summary_df['Row Count'].sum(),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'See log file for details'
                    ]
                }
                
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Format summary file
            self.format_excel_file(summary_file)
            
            logger.info(f"📋 Created summary report: {summary_file.name}")
            
        except Exception as e:
            logger.error(f"❌ Failed to create summary report: {e}")
    
    def close_connection(self):
        """Close database connection"""
        if self.connection:
            self.connection.close()
            logger.info("🔌 Database connection closed")

def main():
    """Main execution function"""
    exporter = None
    try:
        exporter = EliteDatabaseExporter()
        exporter.export_all_tables()
    except Exception as e:
        logger.error(f"❌ Export failed: {e}")
        sys.exit(1)
    finally:
        if exporter:
            exporter.close_connection()

if __name__ == "__main__":
    main()
